import random
from openpyxl import load_workbook ,Workbook
from constants import reelsetSymbolsCount , freesetSymbolsCount , excelFilePath

def build_reel_column(symbol_counts: dict, col_key: int) -> list:
    """Build a shuffled list of symbols for a given column key."""
    reel = []
    for symbol, counts in symbol_counts.items():
        reel.extend([symbol] * counts[col_key])
    random.shuffle(reel)
    return reel

def write_reels_to_excel(symbol_counts: dict):
    wb = load_workbook(excelFilePath)
    if "Reels" in wb.sheetnames:
        ws = wb["Reels"]  # Use existing sheet
    else:
        ws = wb.create_sheet(title="Reels")

    num_cols = len(next(iter(symbol_counts.values())))  # 5 columns

    # Write header row
    for col in range(1, num_cols + 1):
        ws.cell(row=1, column=col, value=f"Reel {col}")

    # Build and write each column
    for col in range(1, num_cols + 1):
        reel = build_reel_column(symbol_counts, col-1)
        for row_idx, symbol in enumerate(reel, start=2):  # start=2 to skip header
            ws.cell(row=row_idx, column=col, value=symbol)

    wb.save(excelFilePath)
    print(f"Saved to {excelFilePath}")

write_reels_to_excel(freesetSymbolsCount)