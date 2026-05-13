from openpyxl import load_workbook
from constants import excelFilePath , GameFeatures , baseReelWorkBook , freeReelWorkBook
import random

def get_col_data(gameFeature: GameFeatures,col: int) -> list:
    wb = load_workbook(excelFilePath, data_only=True)
    match gameFeature:
        case GameFeatures.BASEGAME:
            ws = wb[baseReelWorkBook]
        case GameFeatures.FREEGAME:
            ws = wb[freeReelWorkBook]
    ws = wb.active
    
    return [
        ws.cell(row=row, column=col).value
        for row in range(2, ws.max_row + 1)
        if ws.cell(row=row, column=col).value is not None
    ]

def collect_col_data(gameFeature: GameFeatures):
    reels =[]
    for i in range(1,6):
        data = get_col_data(gameFeature,i)
        reels.append(data)

    return reels

baseReels = collect_col_data(GameFeatures.BASEGAME)
freeReels = collect_col_data(GameFeatures.FREEGAME)

def generate_slot_matrix(gameFeature: GameFeatures) -> list[list[str]]:
    """Generate a 3x5 matrix from 5 reel columns."""
    matrix = []
    
    for col in range(5):  # 5 reels
        match gameFeature:
            case GameFeatures.BASEGAME:
                reel = baseReels[col] 
            case GameFeatures.FREEGAME:
                reel = freeReels[col]  
        
        max_start = len(reel) - 3
        start_idx = random.randint(0, max_start)
        window = reel[start_idx : start_idx + 3]  # 3 consecutive symbols
        
        matrix.append(window)
    
    # matrix is currently 5 cols x 3 rows, transpose to 3 rows x 5 cols
    matrix_3x5 = list(map(list, zip(*matrix)))
    return matrix_3x5

